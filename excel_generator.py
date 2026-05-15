"""
excel_generator.py — Generate a styled Excel workbook from the AI JSON data.
Layout matches the user's Google Sheets template (yellow header / blue period rows / red condition cells).
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ─── Style constants ──────────────────────────────────────────────────────────
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_BLUE   = PatternFill("solid", fgColor="4A86E8")
_RED    = PatternFill("solid", fgColor="CC0000")
_LGRAY  = PatternFill("solid", fgColor="F3F4F6")

_WHITE_BOLD  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_BLACK_BOLD  = Font(bold=True, name="Calibri", size=10)
_BLACK_NORM  = Font(name="Calibri", size=10)
_RED_BOLD_BIG = Font(color="CC0000", bold=True, name="Calibri", size=12)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cell(ws, row, col, value="", fill=None, font=None, align=None, border=None):
    """Write a cell and optionally apply styles. Returns the cell."""
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill      = fill
    if font:   c.font      = font
    if align:  c.alignment = align
    if border: c.border    = border
    return c


def _border_row(ws, row, num_cols=8):
    """Apply thin border to every cell in a row."""
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).border = _THIN


def _clean_price(val):
    """Extract numeric value from strings like '1,500 THB' or '1,500.50'.
    Returns None for non-numeric values like 'N/A' or 'on request only'."""
    if val is None: return None
    if isinstance(val, (int, float)): return float(val)
    # Remove commas and extract first valid float/int pattern
    import re
    s = str(val).replace(',', '').strip()
    if not s: return None
    
    # Detect explicit non-numeric markers
    s_lower = s.lower()
    if s_lower in ('n/a', '-', '--', 'na', 'tba', 'tbd') or 'on request' in s_lower or 'not available' in s_lower:
        return None
    
    # Try to find a number
    match = re.search(r"[-+]?\d*\.\d+|\d+", s)
    if match:
        try: 
            return float(match.group())
        except (ValueError, TypeError): 
            return None
    return None


def generate_comparison_excel(data: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Report"

    # Column widths (A–F)
    widths = [45, 25, 25, 10, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 1
    hotel_name = data.get("hotel_name", "HOTEL NAME")
    year_1     = data.get("year_1", "24/25")
    year_2     = data.get("year_2", "25/26")

    # Header Row
    _cell(ws, row, 1, hotel_name, fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 2, year_1,     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 3, year_2,     fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 4, "%",        fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 5, "",         fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    _cell(ws, row, 6, "",         fill=_YELLOW, font=_BLACK_BOLD, align=_CENTER, border=_THIN)
    row += 1

    seasons = data.get("seasons", [])
    for season in seasons:
        season_name = season.get("season_name") or ""
        p1 = season.get("period_1") or ""
        p2 = season.get("period_2") or ""

        # Period Header (Blue)
        _cell(ws, row, 1, "", fill=_BLUE, border=_THIN)
        _cell(ws, row, 2, p1, fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 3, p2, fill=_BLUE, font=_WHITE_BOLD, align=_CENTER, border=_THIN)
        _cell(ws, row, 4, "", fill=_BLUE, border=_THIN)
        _cell(ws, row, 5, "", fill=_BLUE, border=_THIN)
        _cell(ws, row, 6, season_name, fill=_YELLOW if season_name else _LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        row += 1

        # Conditions for the season can go to Col F on the first room row
        cond_raw = season.get("conditions") or []
        if isinstance(cond_raw, list):
            cond_lines = cond_raw
        else:
            cond_lines = str(cond_raw).split('\n')
        
        rooms = season.get("rooms") or []
        for i, rm in enumerate(rooms):
            name = rm.get("room_name") or ""
            p1_val = rm.get("price_1", 0)
            p2_val = rm.get("price_2", 0)
            
            price_1 = _clean_price(p1_val)
            price_2 = _clean_price(p2_val)

            price_1_disp = price_1 if price_1 is not None else "N/A"
            price_2_disp = price_2 if price_2 is not None else "N/A"

            if price_1 is not None and price_2 is not None:
                diff_amt = price_2 - price_1
                if price_1 > 0:
                    diff_pct = (price_2 - price_1) / price_1
                elif price_1 == 0 and price_2 > 0:
                    diff_pct = 1.0 # 100%
                else:
                    diff_pct = 0.0
                diff_amt_disp = diff_amt
                diff_pct_disp = diff_pct
            else:
                diff_amt = 0
                diff_pct = 0
                diff_amt_disp = "-"
                diff_pct_disp = "-"

            _cell(ws, row, 1, name, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 2, price_1_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            _cell(ws, row, 3, price_2_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if price_1 is not None: ws.cell(row=row, column=2).number_format = '#,##0'
            if price_2 is not None: ws.cell(row=row, column=3).number_format = '#,##0'

            pct_cell = _cell(ws, row, 4, diff_pct_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_pct_disp, float):
                pct_cell.number_format = '0.00%'
                if diff_pct > 0: pct_cell.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_pct < 0: pct_cell.font = Font(color="16A34A", bold=True, name="Calibri", size=10)

            amt_cell = _cell(ws, row, 5, diff_amt_disp, font=_BLACK_NORM, align=_CENTER, border=_THIN)
            if isinstance(diff_amt_disp, float):
                amt_cell.number_format = '+#,##0;-#,##0;0'
                if diff_amt > 0: amt_cell.font = Font(color="CC0000", bold=True, name="Calibri", size=10)
                elif diff_amt < 0: amt_cell.font = Font(color="16A34A", bold=True, name="Calibri", size=10)
            
            cond_text = cond_lines[i] if i < len(cond_lines) else ""
            _cell(ws, row, 6, cond_text, font=Font(color="CC0000", name="Calibri", size=10), align=_LEFT, border=_THIN)
            row += 1

        # If there are remaining conditions, print them in col F with empty room cells
        for i in range(len(rooms), len(cond_lines)):
            if not cond_lines[i].strip(): continue
            _cell(ws, row, 1, "", border=_THIN)
            _cell(ws, row, 2, "", border=_THIN)
            _cell(ws, row, 3, "", border=_THIN)
            _cell(ws, row, 4, "", border=_THIN)
            _cell(ws, row, 5, "", border=_THIN)
            _cell(ws, row, 6, cond_lines[i], font=Font(color="CC0000", name="Calibri", size=10), align=_LEFT, border=_THIN)
            row += 1

    row += 1

    def _extract_notes(lines):
        """Returns (clean_text, notes_list)"""
        if lines is None: return "", []
        if not isinstance(lines, list):
            lines = [str(lines)]
        clean = []
        notes = []
        for line in lines:
            if not line: continue
            line_str = str(line)
            if "[RED]" in line_str:
                notes.append(line_str.replace("[RED]", "").strip())
            else:
                clean.append(line_str)
        return "\n".join(clean), notes

    def _build_stacked_section(title, key):
        nonlocal row
        items = data.get(key, [])
        if not items: return
        
        _cell(ws, row, 1, title, fill=_YELLOW, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        for item in items:
            c1_raw = item.get("contract_1", [])
            c2_raw = item.get("contract_2", [])
            
            c1_text, c1_notes = _extract_notes(c1_raw)
            c2_text, c2_notes = _extract_notes(c2_raw)
            all_notes = c1_notes + [n for n in c2_notes if n not in c1_notes]
            
            # Contract 1
            _cell(ws, row, 1, f"Contract {year_1}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            
            _cell(ws, row, 1, c1_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = max(18, 15 * (c1_text.count("\n") + 1))
            row += 1
            
            # Contract 2
            _cell(ws, row, 1, f"Contract {year_2}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            row += 1
            
            _cell(ws, row, 1, c2_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = max(18, 15 * (c2_text.count("\n") + 1))
            row += 1
            
            # Notes
            if all_notes:
                notes_text = "\n".join(all_notes)
                _cell(ws, row, 1, notes_text, font=Font(color="CC0000", bold=True, name="Calibri", size=10), align=_LEFT, border=_THIN)
                for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.row_dimensions[row].height = max(18, 15 * (notes_text.count("\n") + 1))
                row += 1
                
        row += 1

    def _build_sidebyside_section(title, key):
        nonlocal row
        items = data.get(key, [])
        if not items: return
        
        _cell(ws, row, 1, title, fill=_YELLOW, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        
        # Headers
        _cell(ws, row, 1, f"Contract {year_1}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 2, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        
        _cell(ws, row, 3, "Comparison / Changes", fill=PatternFill("solid", fgColor="D0E0F0"), font=Font(color="000080", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
        _cell(ws, row, 4, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

        _cell(ws, row, 5, f"Contract {year_2}", fill=_LGRAY, font=_BLACK_BOLD, align=_LEFT, border=_THIN)
        _cell(ws, row, 6, "", border=_THIN)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        row += 1

        for item in items:
            c1_raw = item.get("contract_1", [])
            c2_raw = item.get("contract_2", [])
            diff = str(item.get("diff_summary") or "SAME")
            
            c1_text, c1_notes = _extract_notes(c1_raw)
            c2_text, c2_notes = _extract_notes(c2_raw)
            all_notes = c1_notes + [n for n in c2_notes if n not in c1_notes]

            _cell(ws, row, 1, c1_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 2, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

            _cell(ws, row, 3, diff, font=Font(color="000080", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
            _cell(ws, row, 4, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)

            _cell(ws, row, 5, c2_text, font=_BLACK_NORM, align=_LEFT, border=_THIN)
            _cell(ws, row, 6, "", border=_THIN)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            
            ws.row_dimensions[row].height = max(25, 15 * (max(c1_text.count("\n"), c2_text.count("\n")) + 1))
            row += 1
            
            # Notes span the full width below the side-by-side cells
            if all_notes:
                notes_text = "\n".join(all_notes)
                _cell(ws, row, 1, notes_text, font=Font(color="CC0000", bold=True, name="Calibri", size=10), align=_CENTER, border=_THIN)
                for c in range(2, 7): _cell(ws, row, c, "", border=_THIN)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.row_dimensions[row].height = max(18, 15 * (notes_text.count("\n") + 1))
                row += 1
                
        row += 1

    # Extra Bed is stacked in the image
    _build_stacked_section("Extra bed / Extra person", "extra_bed")
    
    # Early Bird, Bonus Night, Wellbeing, Cancellation are side-by-side
    _build_sidebyside_section("Early Bird Offer", "early_bird")
    _build_sidebyside_section("BONUS NIGHT OFFER", "bonus_night")
    _build_sidebyside_section("WELLBEING SANCTUARY POOL SUITE LONG STAY BENEFITS", "wellbeing")
    _build_sidebyside_section("Cancellation", "cancellation")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
